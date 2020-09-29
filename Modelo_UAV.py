# -*- coding: utf-8 -*-
"""
Created on Mon Aug 10 12:16:35 2020

@author: Karol
"""

from ortools.linear_solver import pywraplp
import numpy as np
import openpyxl
import math
from collections import defaultdict

import endurance_calculator
import distance_functions

Metros_por_milla= 1609.34

TYPE_TRUCK 		= 1
TYPE_UAV 		= 2

def make_dict():
	return defaultdict(make_dict)

class make_node:
	def __init__(self, nodeType, latDeg, lonDeg, altMeters, parcelWtLbs, serviceTimeTruck, serviceTimeUAV, address):
		# Set node[nodeID]
		self.nodeType 			= nodeType
		self.latDeg 			= latDeg
		self.lonDeg				= lonDeg
		self.altMeters			= altMeters
		self.parcelWtLbs 		= parcelWtLbs
		self.serviceTimeTruck	= serviceTimeTruck	# [seconds]
		self.serviceTimeUAV 	= serviceTimeUAV	# [seconds]
		self.address 			= address			# Might be None

class make_vehicle:
	def __init__(self, vehicleType, takeoffSpeed, cruiseSpeed, landingSpeed, yawRateDeg, cruiseAlt, capacityLbs, launchTime, recoveryTime, serviceTime, batteryPower, flightRange):
		# Set vehicle[vehicleID]
		self.vehicleType	= vehicleType
		self.takeoffSpeed	= takeoffSpeed
		self.cruiseSpeed	= cruiseSpeed
		self.landingSpeed	= landingSpeed
		self.yawRateDeg		= yawRateDeg
		self.cruiseAlt		= cruiseAlt
		self.capacityLbs	= capacityLbs
		self.launchTime		= launchTime	# [seconds].
		self.recoveryTime	= recoveryTime	# [seconds].
		self.serviceTime	= serviceTime
		self.batteryPower	= batteryPower	# [joules].
		self.flightRange	= flightRange	# 'high' or 'low'

class make_travel:
	def __init__(self, takeoffTime, flyTime, landTime, totalTime, takeoffDistance, flyDistance, landDistance, totalDistance):
		# Set travel[vehicleID][fromID][toID]
		self.takeoffTime 	 = takeoffTime
		self.flyTime 		 = flyTime
		self.landTime 		 = landTime
		self.totalTime 		 = totalTime
		self.takeoffDistance = takeoffDistance
		self.flyDistance	 = flyDistance
		self.landDistance	 = landDistance
		self.totalDistance	 = totalDistance
        
def solve_mfstsp_IP(node, vehicle, travel, cutoffTime, Etype, V, N_0, C, Pp, Nmas, Cnum, tau_prima, eee):
    p=[]
   
    for v in V:
        for i in N_0:
            for j in C:
                if ((j != i) and (Pp[j] <= 5)):
                   for k in Nmas:
                       if (k != i) and (k != j):							
                           # Calculate the endurance for each sortie:
                           if (k == Cnum+1):
                               eee[v][i][j][k] = endurance_calculator.give_endurance(node, vehicle, travel, v, i, j, 0, Etype)
                           else:
                               eee[v][i][j][k] = endurance_calculator.give_endurance(node, vehicle, travel, v, i, j, k, Etype)
                               # If endurance is based on distance, build the P set using distance limitations:
                               if Etype == 5:
                                   DISTij = distance_functions.groundDistanceStraight(node[i].latDeg*(math.pi/180), node[i].lonDeg*(math.pi/180), node[j].latDeg*(math.pi/180), node[j].lonDeg*(math.pi/180))
                                   DISTjk = distance_functions.groundDistanceStraight(node[j].latDeg*(math.pi/180), node[j].lonDeg*(math.pi/180), node[k].latDeg*(math.pi/180), node[k].lonDeg*(math.pi/180))
                                   if vehicle[v].flightRange == 'low':
                                       if DISTij + DISTjk <= 6*Metros_por_milla:
                                           p.append([v,i,j,k])
                                       elif vehicle[v].flightRange == 'high':
                                           if DISTij + DISTjk <= 12*Metros_por_milla:
                                               p.append([v,i,j,k])
                                   else:
                                       if (tau_prima[v][i][j] + node[j].serviceTimeUAV + tau_prima[v][j][k] <= eee[v][i][j][k]):
                                           p.append([v,i,j,k])
    return p

def main ():
    solver = pywraplp.Solver ( 'SolveAssignmentProblemMIP' , 
                           pywraplp.Solver.CBC_MIXED_INTEGER_PROGRAMMING ) 
    
    
 
    wb = openpyxl.load_workbook('instancias.xlsx')
    wshoja = wb['Hoja1']
    Cnum=int(wshoja.cell(row=1, column=2).value)
    numUAVs=int(wshoja.cell(row=1, column=4).value)
    node={}
    vehicle={}
    travel=defaultdict(make_dict)
    for i in range(0,Cnum):
        nodeID 				= int(wshoja.cell(row=3+i, column=1).value)
        nodeType			= int(wshoja.cell(row=3+i, column=2).value)
        latDeg				= float(wshoja.cell(row=3+i, column=3).value)
        lonDeg		    	= float(wshoja.cell(row=3+i, column=4).value)
        altMeters			= float(wshoja.cell(row=3+i, column=5).value)
        parcelWtLbs		    = float(wshoja.cell(row=3+i, column=3).value)
        node[nodeID] = make_node(nodeType, latDeg, lonDeg, altMeters, parcelWtLbs)
    
    
    tmpUAVs = 0
    for i in range(0,Cnum):
        vehicleID 			= int(wshoja.cell(row=((Cnum+1)**2+(Cnum+10)+i), column=1).value)
        vehicleType			= int(wshoja.cell(row=((Cnum+1)**2+(Cnum+10)+i), column=2).value)
        takeoffSpeed		= float(wshoja.cell(row=((Cnum+1)**2+(Cnum+10)+i), column=3).value)
        cruiseSpeed			= float(wshoja.cell(row=((Cnum+1)**2+(Cnum+10)+i), column=4).value)
        landingSpeed		= float(wshoja.cell(row=((Cnum+1)**2+(Cnum+10)+i), column=5).value)
        yawRateDeg			= float(wshoja.cell(row=((Cnum+1)**2+(Cnum+10)+i), column=6).value)
        cruiseAlt			= float(wshoja.cell(row=((Cnum+1)**2+(Cnum+10)+i), column=7).value)
        capacityLbs			= float(wshoja.cell(row=((Cnum+1)**2+(Cnum+10)+i), column=8).value)
        launchTime			= float(wshoja.cell(row=((Cnum+1)**2+(Cnum+10)+i), column=9).value)
        recoveryTime		= float(wshoja.cell(row=((Cnum+1)**2+(Cnum+10)+i), column=10).value)
        serviceTime			= float(wshoja.cell(row=((Cnum+1)**2+(Cnum+10)+i), column=11).value)
        batteryPower		= float(wshoja.cell(row=((Cnum+1)**2+(Cnum+10)+i), column=12).value)
        flightRange			= str(wshoja.cell(row=((Cnum+1)**2+(Cnum+10)+i), column=13).value)
			
        if (vehicleType == TYPE_UAV):
            tmpUAVs += 1
            if (tmpUAVs <= numUAVs):
                vehicle[vehicleID] = make_vehicle(vehicleType, takeoffSpeed, cruiseSpeed, landingSpeed, yawRateDeg, cruiseAlt, capacityLbs, launchTime, recoveryTime, serviceTime, batteryPower, flightRange)
        else:
            vehicle[vehicleID] = make_vehicle(vehicleType, takeoffSpeed, cruiseSpeed, landingSpeed, yawRateDeg, cruiseAlt, capacityLbs, launchTime, recoveryTime, serviceTime, batteryPower, flightRange)
    
    for i in range(0,Cnum):	
        tmpi 	= int(wshoja.cell(row=(Cnum+6)+i, column= 1)).value	
        tmpj 	= int(wshoja.cell(row=(Cnum+6)+i, column= 2)).value
        tmpTime	= float(wshoja.cell(row=(Cnum+6)+i, column= 3)).value
        tmpDist	= float(wshoja.cell(row=(Cnum+6)+i, column= 4)).value
        for vehicleID in vehicle:
            if (vehicle[vehicleID].vehicleType == TYPE_TRUCK):
                travel[vehicleID][tmpi][tmpj] = make_travel(0.0, tmpTime, 0.0, tmpTime, 0.0, tmpDist, 0.0, tmpDist)
    
    
    
    C=np.zeros((Cnum,1))
    N=np.zeros((Cnum+1,1))
    N_0=np.zeros(((Cnum+1)**2,1))
    Nmas=np.zeros(((Cnum+1)**2,1))
    tau=np.zeros(((Cnum+1)**2,1))
    tau_prima=np.zeros(((Cnum+1)**2,1))
    Pp=np.zeros((Cnum+1,1))
    C_v=np.zeros((Cnum+1,1))
    
    V=int(wshoja.cell(row=1, column=4).value)
    
    for i in range(Cnum):
        C[i]= int(wshoja.cell(row=4+i, column=1).value)
    
    for i in range (Cnum+1):
        N[i]=int(wshoja.cell(row=3+i, column=1).value)        
    
    for i in range((Cnum+1)**2):
        N_0[i] = wshoja.cell(row=(Cnum+6)+i, column= 1).value
        Nmas[i] = wshoja.cell(row=(Cnum+6)+i, column= 2).value
        tau[i] = wshoja.cell(row=(Cnum+6)+i, column= 3).value
        tau_prima[i] = wshoja.cell(row=(Cnum+6)+i, column= 3).value
    
            
    sL = int(wshoja.cell(row=((Cnum+1)**2+(Cnum+10)), column=9).value)
        
    sR = int(wshoja.cell(row=((Cnum+1)**2+(Cnum+10)), column=10).value)
    
    sigma = int(wshoja.cell(row=((Cnum+1)**2+(Cnum+10)), column=11).value)
    
    sigma_prima = int(wshoja.cell(row=((Cnum+1)**2+(Cnum+10)), column=11).value)
    
    for i in range(Cnum+1):
        Pp[i]=int(wshoja.cell(row=3+i, column=6).value)
        C_v=np.where(Pp<=5)
    
    eee = defaultdict(make_dict)
    
    p = solve_mfstsp_IP(node, vehicle, travel, cutoffTime, Etype, V, N_0, C, Pp, Nmas, Cnum, tau_prima,eee)

    M = int(wshoja.cell(row=1, column=6).value)
        
    
    #Variables de decisión
           
    x = {}
    for i in  (N_0):
        for j in  (Nmas):
            x[i,j] = solver.BoolVar('x[i%i,j%i]' % (i,j))
            
    y = {}
    for v in V:
        for i in  (N_0):
            for j in C_v:
                for k in (Nmas):
                    y[v,i,j,k] = solver.BoolVar('y[v%i,i%i,j%i,k%i]'% (v,i,j,k))
                    
    p = {}
    for i in  (N_0):
        for j in C:
            p[i,j] = solver.BoolVar('p[i%i,j%i]' % (i,j))
            
       #Variables Camión
            
    t_arriba = {}
    for i in  (N):
        t_arriba[i] = solver.NumVar(0,solver.Infinity(),'t_arriba[i%i]' %(i))
        
    t_abajo = {}
    for i in  (Nmas):
         t_abajo[i] = solver.NumVar(0,solver.Infinity(),'t_abajo[i%i]' %(i))
    
    t_barra = {}
    for i in  (N):
         t_barra[i] = solver.NumVar(0,solver.Infinity(),'t_barra[i%i]' %(i))
         
       # Variables UAV
         
    t_arriba_prima = {}
    for v in V:
        for i in  (N):
             t_arriba_prima[v,i] = solver.NumVar(0,solver.infinity(),'t_arriba_prima[v%i,i%i]' %(v,i))
    
    t_abajo_prima = {}
    for v in V:
        for i in  (N):
             t_abajo_prima[v,i] = solver.NumVar(0,solver.infinity(),'t_abajo_prima[v%i,i%i]' %(v,i))
    
      #--------------------
             
    zR = {}
    for v_1 in V:
        for v_2 in V:
            if v_2 != v_1:
                for k in  (Nmas):
                    zR[v_1,v_2,k] = solver.BoolVar('zR[v_1%iv_2%ik%i]' % (v_1,v_2,k))
                    
    for k in  (Nmas):
        for v in V:
            zR[v,k] = solver.BoolVar('zR[v%ik%i]' % (v,k))
            #zR[0,v,k] = solver.BoolVar('zR[v%ik%i]' % (0,v,k))
                    
    for v in V:
        for k in (Nmas):
            zR[v,k] = solver.BoolVar('zR[v%ik%i]' % (v,k))
            #zR[v,0,k] = solver.BoolVar('zR[v%ik%i]' % (v,0,k))
    
    zL = {}
    for v_1 in V:
        for i in (N_0):
            for v_2 in V:
                if v_2 != v_1:
                    zL[v_1,v_2,i] = solver.BoolVar('zL[v_1%iv_2%ii%i]' % (v_1,v_2,i))
    
    for v in V:
        for i in (N_0):
                zL[v,i] = solver.BoolVar('zL[v%ii%i]' % (v,i))
                #zL[0,v,i] = solver.BoolVar('zL[v%ii%i]' % (0,v,i))
    
    for v in V:
        for i in (N_0):
            zL[v,i] = solver.BoolVar('zL[v%ii%i]' % (v,i))
            #zL[v,0,i] = solver.BoolVar('zL[v%ii%i]' % (v,0,i))
                    
    z_prima = {}
    for i in C:
        for v_1 in V:
            for v_2 in V:
                z_prima[i,v_1,v_2] = solver.BoolVar('z_prima[v%iv_1%iv_2%i]' % (i,v_1,v_2))
                
    z_comillas = {}
    for i in C:
        for v_1 in V:
            for v_2 in V:
                z_comillas[i,v_1,v_2] = solver.BoolVar('z_comillas[i%iv_1%iv_2%i]' % (i,v_1,v_2))
                
    u = {}
    for i in (Nmas):
        u[i] = solver.NumVar(1,C+2,'i%i' %(i))
    
                            
    #1 Función Obj
    
    solver.Minimize(t_abajo[(C+1)])
    
    #Restricciones
    
    # 2
    for j in C:
        solver.add(solver.sum([x[(i,j)] for i in N_0 if i!=j])+(solver.sum([y[(v,i,j,k)] for v in V for i in N_0 if i != j for k in Nmas]))==1)  
    
    # 3
    solver.Add(solver.Sum([x[(0,j)] for j in Nmas]) == 1)
    
    # 4
    solver.Add(solver.Sum([x[(i,C+1)] for i in N_0]) == 1) 
    
    # 5
    for j in C:
        solver.add(solver.sum([x[(i,j)] for i in N_0 if i !=j]) == (solver.sum ([x[(j,k)] for k in Nmas if k != j ])))
        
    # 6
    for i in N_0:
        for v in V:
            solver.add(solver.sum([y[(v,i,j,k)] for j in C if j != i for k in Nmas]) <= 1)
    
    # 7
    for k in Nmas:
        for v in V:
            solver.add(solver.sum([y[(v,i,j,k)] for i in N_0 if i != k for j in C]) <= 1)
            
    # 8
    for v in V :
        for i in C:
            for j in C:
                if j != i:
                    for k in Nmas:
                        solver.add(2*y[(v,i,j,k)] <= solver.sum([x[(h,i)] for h in N_0 if h != i]) + solver.sum([x[(l,k)] for l in C if l != k]))
                    
    # 9
    for v in V:
        for j in C:
            for k in Nmas:
                for i in C:
                    if i == 0:
                        solver.add([y[(v,0,j,k)]] <= solver.sum([x[(h,k)] for h in N_0 if h!=k]))
                
    # 10
    for i in C:
        for k in Nmas:
            if k != i:
                for v in V:
                    solver.add([u[(k)]-u[(i)]] >= (1-(C+2))(1-solver.sum([1-y[(v,i,j,k)] for j in C])))
                    
    # 11
    for i in C:
        for j in Nmas:
            if j != i:
                solver.add([u[(i)]-u[(j)]+1] <= (C+2)(1-[x[(i,j)]]))
                
    # 12
    for i in C:
        for j in N:
            if j != i:
                solver.add([u[(i)]-u[(j)]] >= (1-(C+2)[p[(i,j)]]))
                         
    # 13
    for i in C:
        for j in C:
            if j != i:
                solver.add(u[(i)]-u[(j)] <= (-1+(C+2)(1-p[(i,j)]))) 
                           
    # 14
    for i in C:
        for j in C:
            if j != i:
                solver.add([p[(i,j)] + p[(i,j)]]==1)
                
    # 15
    for i in N_0:
        for k in Nmas:
            if k != i:
                for l in C:
                    if l != i and l != k:
                        for v in V:
                           solver.add(t_abajo_prima[(v,l)] >= t_arriba_prima[(v,k)] - M*(3 - solver.sum([y[(v,i,j,k)] for j in C if j != l]) - solver.sum([y[(v,l,m,n)] for m in C if (m !=i and m !=k and m !=l) for n in Nmas if (n !=i and n !=k)])-p[(i,l)]))
                            
    # 16
    for v in V:
        for i in N_0:
            solver.add(t_abajo_prima[(v,i)] >= t_arriba_prima[(v,i)] + sL[(v,i)] - M*(1-solver.sum([y[(v,i,j,k)] for j in C if j != i for k in Nmas])))
            
    #17
    for v in V:
        for i in N_0:
            solver.add(t_abajo_prima[(v,i)] >= t_arriba[(i)] + sL[(v,i)] - M*(1-zL[(v,0,i)]))
            
    #18
    for v in V:
        for i in N_0:
           solver.add(t_abajo_prima[(v,i)] >= t_barra[(i)] + sL[(v,i)] - M*(1-zL[(v,0,i)]))
            
    #19
    for v in V:
        for v_2 in V:
            if v_2 != v:
                for i in N_0:
                   solver.add(t_abajo_prima[(v,i)] >= t_abajo_prima[(v_2,i)] + sL[(v,i)] - M*(1-zL[(v_2,v,i)]))
    
    #20
    for v in V:
        for v_2 in V:
            if v_2 != v:
                for i in C:
                    solver.add(t_abajo_prima[(v_2,i)] >= t_arriba_prima[(v_2,i)] + sL[(v,i)] - M*(1-zL[(v_2,v,i)]))
                    
    #21
    for v in V:
        for j in C:
            for i in N_0:
                if i != j:
                    solver.add(t_arriba_prima[(v,j)] >= t_abajo_prima[(v,i)] + tau_prima[(v,i,j)] - M*(1-solver.sum([y[(v,i,j,k)] for k in Nmas])))
                    
    #22
    for v in V:
        for j in C:
            for i in N_0:
                if i != j:
                    solver.add(t_arriba_prima[(v,j)] <= t_abajo_prima[(v,i)] + tau_prima[(v,i,j)] + M*(1-solver.sum([y[(v,i,j,k)] for k in Nmas])))
                    
    #23
    for v in V:
        for j in C:
            solver.add(t_abajo_prima[(v,j)] >= t_arriba_prima[(v,j)] + sigma_prima[(v,j)] (solver.sum([y[(v,i,j,k)] for i in N_0 if i != j for k in Nmas])))
            
    #24
    for v in V:
        for j in C:
           solver.add(t_abajo_prima[(v,j)] <= t_arriba_prima[(v,j)] + sigma_prima[(v,j)] + M*(1-solver.sum([y[(v,i,j,k)] for i in N_0 if i != j for k in Nmas])))
             
    #25
    for v in V:
        for k in Nmas:
            solver.add(t_arriba_prima[(v,k)] >= t_arriba[(k)] + sR[(v,k)] - M*(1-zR[(v,0,k)]))
            
    #26
    for v in V:
        for k in Nmas:
            solver.add(t_arriba_prima[(v,k)] >= t_barra[(k)] + sR[(v,k)] - M*(1-zR[(0,v,k)]))
            
    #27
    for v in V:
        for v_2 in V:
            if v_2 != v:
                for k in Nmas:
                    solver.add(t_arriba_prima[(v,k)] >= t_arriba[(v_2,k)] + sR[(v,k)] - M*(1-zR[(v_2,v,k)]))
                    
    #28
    for v in V:
        for v_2 in V:
            if v_2 != v:
                for k in C:
                   solver.add(t_arriba_prima[(v,k)] >= t_arriba[(v_2,k)] + sR[(v,k)] - M*(1-z_prima[(v_2,v,k)]))
                    
    #29
    for v in V:
        for k in Nmas:
            for i in C:
                if j != k:
                    solver.add(t_arriba_prima[(v,k)] >= t_abajo_prima[(v,j)] + tau_prima[(v,j,k)] + sR[(v,k)] - M*(1-solver.sum([y[(v,i,j,k)] for i in N_0])))
                    
    #30
    for v in V:
        for i in N_0:
            for j in C_v:
                if j != i:
                    for k in Nmas:
                      solver.add((t_arriba_prima[(v,k)] - sR[(v,k)]) - t_abajo_prima[(v,i)] <= eee[(v,i,j,k)] + M*(1-y[(v,i,j,k)]))
                        
    # 31
    for i in N_0:
        for j in Nmas:
            if j != i:
                solver.Add(t_arriba[(i)] >= t_abajo[(i)] + tau[(i,j)] - M*(1 - x[(i,j)]))
                
    # 32
    for k in Nmas:
        solver.add(t_barra[(k)] >= t_arriba[(k)] + sigma[(k)](solver.sum([x[(j,k)] for j in N_0 if j != k])))
    
    # 33
    for k in Nmas:
        for v in V:
            solver.add(t_barra[(k)] >= t_arriba_prima[(v,k)] + sigma[(k)] - M*(1 - zR[(0,v,k)]))
    
    # 34
    for k in C:
        for v in V:
           solver.add(t_barra[(k)] >= t_abajo_prima[(v,k)] + sigma[(k)] - M*(zL[(v,0,k)]))
                            
    # 35
    for k in Nmas:
      solver.add(t_abajo[(k)] >= t_barra[(k)])
    
    # 36
    for k in Nmas:
        for v in V:
            solver.add(t_abajo[(k)] >= t_arriba_prima[(v,k)] - M*(1-solver.sum([y[(v,i,j,k)] for i in N_0 if i!= k for j in C])))
        
    # 37
    for k in N_0:
        for v in V:
            solver.add(t_abajo[(k)] >= t_abajo_prima[(v,k)] - M*(1-solver.sum([y[(v,k,l,m)] for l in C if l!= k for m in Nmas])))
    
    # 38
    for v in V:
        for k in Nmas:
            solver.add(zR[(0,v,k)] + zR[(v,0,k)] == solver.sum([y[(v,j,i,k)] for i in N_0 if i != k for j in C]))
    
    # 39
    for v in V:
        for v_2 in V:
            if v_2 != v:
                for k in Nmas:
                    solver.add(zR[(v,v_2,k)] <= solver.sum([y[(v,i,j,k)] for i in N_0 if i != k for j in C]))
    # 40
    for v in V:
        for v_2 in V:
            if v_2 != v:
                for k in Nmas:
                   solver.add(zR[(v,v_2,k)] <= solver.sum([y[(v_2,i,j,k)] for i in N_0 if i != k for j in C]))
    
    # 41
    for v in V:
        for v_2 in V:
            if v_2 != v:
                for k in Nmas:
                    solver.add(zR[(v,v_2,k)] + zR[(v_2,v,k)] <= 1)    
    
    # 42
    for v in V:
        for v_2 in V:
            if v_2 != V:
                for k in Nmas:
                    solver.add(zR[(v,v_2,k)] + zR[(v_2,v,k)] + 1 >= solver.sum([y[(v,i,j,k)] for i in N_0 if i != k for j in C]))
                    
                        
    # 43
    for v in V:
        for i in N_0:
            solver.add(zL[(0,v,i)] + zL[(v,0,i)] == solver.sum([y[(v,i,j,k)] for j in C if j != i for k in Nmas]))
            
    # 44
    for v in V:
        for v_2 in V:
            if v_2 != v:
                for i in N_0:
                    solver.add(zL[(v,v_2,i)] <= solver.sum([y[(v,i,j,k)] for j in C if j != i for k in Nmas]) + solver.sum([y[(v_2,i,j,k)] for i in N_0 if i != k for j in C]))
                    
    # 45
    for v in V:
        for v_2 in V:
            if v_2 != v:
                for i in N_0:
                   solver.add(zL[(v,v_2,i)] <= solver.sum([y[(v_2,i,j,k)] for j in C if j != i for k in Nmas]))
                    
    # 46
    for v in V:
        for v_2 in V:
            if v_2 != v:
                for i in N_0:
                  solver.add(zL[(v,v_2,i)] + zL[(v_2,v,i)] <= 1)
                    
    # 47
    for v in V:
        for v_2 in V:
            if v_2 != v:
                for i in N_0:
                    solver.add(zL[(v,v_2,i)] + zL[(v_2,v,i)] + 1 >= solver.sum([y[(v,i,j,k)] for j in C if j != i for k in Nmas]) + solver.sum([y(v_2,i,j,k)] for j in C if j != i for k in Nmas))
                     
    # 48
    for v_2 in V:
        for v in V:
            if v != v_2:
                for k in C:
                  solver.add(z_prima[(v_2,v,k)] <= solver.sum([y[(v_2,k,l,m)] for l in C if l != k for m in Nmas]))
                    
    # 49
    for v in V:
        for v_2 in V:
            if v_2 != v:
                for k in C:
                   solver.add(z_comillas[(v_2,v,k)] <= solver.sum([y[(v,k,l,m)] for l in C if l != k for m in Nmas]))
                                        
    # 50
    for v in V:
        for v_2 in V:
            if v_2 != v:
                for k in C:
                   solver.add(z_prima[(v_2,v,k)] <= solver.sum([y[(v,i,j,k)] for i in N_0 if i != k for j in C]))
                    
    # 51
    for v_2 in V:
        for v in V:
            if v != v_2:
                for k in C:
                    solver.add(z_comillas[(v_2,v,k)] <= solver.sum([y[(v_2,i,j,k)] for i in N_0 if i != k for j in C]))
                    
    # 52
    for v in V:
        for v_2 in V:
            if v_2 != v:
                for k in C:
                   solver.add(z_prima[(v_2,v,k)] + z_comillas[(v,v_2,k)] + 1 >= solver.sum([y[(v,i,j,k)] for i in N_0 if i != k for j in C]) + solver.sum([y[(v_2,i,j,k)] for l in C if l != k for m in Nmas]))
                    
    # 53
    for v in V:
        for v_2 in V:
            if v_2 != v:
                for k in C:
                    solver.add(z_prima[(v_2,v,k)] + z_comillas[(v,v_2,k)] <= 1)
                    
    # 54
    for v in V:
        for v_2 in V:
            if v_2 != v:
                for k in C:
                    solver.add(z_prima[(v_2,v,k)] + z_prima[(v,v_2,k)] <= 1)
                    
    # 55
    for v in V:
        for v_2 in V:
            if v_2 != v:
                for k in C:
                    solver.add(z_comillas[(v_2,v,k)] + z_comillas[(v,v_2,k)] <= 1)
                    
    #correr el modelo
    status = solver.Solve()
    print(status)
    #0:  Optimo
    #1: Factible
    #2: Infactible
    #3: No acotado
    if status == pywraplp.Solver.OPTIMAL or status == pywraplp.Solver.FEASIBLE:
        assert solver.VerifySolution(1e-7, True)
        guardar( x, y , p, t_arriba, t_abajo, t_barra, t_arriba_prima, t_abajo_prima, solver, zR, zL, u, z_comillas, z_prima, Nmas, N_0, C_v, V, N, C)
        if status == pywraplp.Solver.OPTIMAL:
            print('Optimal solution found')
        else:
            print('Feasible solution found')

    else:
        print('Problem not feasible or unbounded')                
  
    
    # Exportar a excel
def guardar( x, y , p, t_arriba, t_abajo, t_barra, t_arriba_prima, t_abajo_prima, solver, zR, zL, u, z_comillas, z_prima, Nmas, N_0, C_v, V, N, C):
    wb = openpyxl.load_workbook('Solucion.xlsx')
    wshoja= wb['hoja']
    fila=1
    for i in N_0:
        for j in Nmas:
            fila=fila+1
            wshoja.cell(row = fila , column= 1).value = 'i'+ str(i)+'j'+ str(j)
            wshoja.cell(row = fila , column= 2).value = x[i,j].solution_value()
    fila2=1        
    for v in V:
        for i in N_0:
            for j in C_v:
                for k in Nmas:
                    fila2=fila2+1
                    wshoja.cell(row = fila2 , column= 3).value = 'v'+ str(v)+'i'+ str(i)+'j'+str(j)+'k'+str(k) 
                    wshoja.cell(row = fila2 , column= 4).value = y[v,i,j,k].solution_value()
                   
    fila3=1
    for i in N_0:
        for j in C:
            fila3=fila3+1
            wshoja.cell(row = fila3 , column= 5).value = 'i'+ str(i)+'j'+ str(j)
            wshoja.cell(row = fila3 , column= 6).value = p[i,j].solution_value()
            
    for i in N:
        wshoja.cell(row = i+1 , column= 7).value = 'i'+ str(i)
        wshoja.cell(row = i+1 , column= 8).value = t_arriba[i].solution_value()
        
    for i in Nmas:
        wshoja.cell(row = i+1 , column= 9).value = 'i'+ str(i)
        wshoja.cell(row = i+1 , column= 10).value = t_abajo[i].solution_value()
         
    for i in N:
        wshoja.cell(row = i+1 , column= 11).value = 'i'+ str(i)
        wshoja.cell(row = i+1 , column= 12).value = t_barra[i].solution_value()
        
    fila4=1
    for v in V:
        for i in N:
            fila4=fila4+1
            wshoja.cell(row = fila4 , column= 13).value = 'v'+ str(v)+'i'+str(i)
            wshoja.cell(row = fila4 , column= 14).value = t_arriba_prima[v,i].solution_value()
            
    fila5=1
    for v in V:
        for i in N:
            fila5=fila5+1
            wshoja.cell(row = fila5 , column= 15).value = 'v'+ str(v)+'i'+str(i)
            wshoja.cell(row = fila5 , column= 16).value = t_abajo_prima[v,i].solution_value()
            
    fila6=1
    for v_1 in V:
        for v_2 in V:
            if v_2 != v_1:
                for k in Nmas:
                    fila6=fila6+1
                    wshoja.cell(row = fila6 , column= 17).value = 'v_1'+ str(v_1)+'v_2'+str(v_2)+'k'+str(k)
                    wshoja.cell(row = fila6 , column= 18).value = zR[v_1,v_2,k].solution_value()
    fila7=1               
    for k in Nmas:
        for v in V:
            fila7=fila7+1
            wshoja.cell(row = fila7 , column= 17).value = 'v_1'+ str(v_1)+'v_2'+str(v_2)+'k'+str(k)
            wshoja.cell(row = fila7 , column= 18).value = zR[v_1,v_2,k].solution_value()
            
    
    fila8=1
    for v_1 in V:
        for i in N_0:
            for v_2 in V:
                if v_2 != v_1:
                    fila8=fila8+1
                    wshoja.cell(row = fila8 , column= 17).value = 'v_1'+ str(v_1)+'v_2'+str(v_2)+'i'+str(i)
                    wshoja.cell(row = fila8 , column= 18).value = zL[v_1,v_2,i].solution_value()
    fila9=1 
    for i in C:
        for v_1 in V:
            for v_2 in V:
                fila9=fila9+1
                wshoja.cell(row = fila9 , column= 19).value = 'v_1'+ str(v_1)+'v_2'+str(v_2)+'i'+str(i)
                wshoja.cell(row = fila9 , column= 20).value = z_prima[v_1,v_2,i].solution_value()
                
    fila10=1
    for i in C:
        for v_1 in V:
            for v_2 in V:
                fila10=fila10+1
                wshoja.cell(row = fila10 , column= 21).value = 'v_1'+ str(v_1)+'v_2'+str(v_2)+'i'+str(i)
                wshoja.cell(row = fila10 , column= 22).value = z_comillas[v_1,v_2,i].solution_value()
                
    for i in Nmas:
        wshoja.cell(row = i+1 , column= 23).value = 'i'+ str(i)
        wshoja.cell(row = i+1 , column= 24).value = u[i].solution_value()
                 
        
        
    
    wshoja.cell(row = 2 , column= 26).value = solver.Objective().Value()
    wb.save('Solucion.xlsx')
   
if __name__ == '__main__':
    main()
   
 
                    
    
                    
    
                    
        